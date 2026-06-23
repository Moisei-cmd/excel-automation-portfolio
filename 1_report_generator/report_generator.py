from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import pandas as pd
import random
from datetime import timedelta, datetime
import os


#  --- генерируем сырой выдуманный отчет ---
prices = {'Механическая клавиатура': 8000, 'Мембранная клавиатура': 2000, 'Игровая мышь': 5000, 'Офисная мышь': 1000}

data_list = []

for day in range(30):
    try:
        curent_date = datetime(2026,6,1) + timedelta(days=day)
        date_str = curent_date.strftime('%d.%m.%Y')
        num_sales = random.randint(1, 4)
        for _ in range(num_sales):
            product = random.choice(list(prices.keys()))
            price = prices[product]
            qty = random.randint(1, 3)
            record = {'Дата': date_str, 'Товар': product, 'Цена': price, 'Количество': qty}
            data_list.append(record)
    except Exception as e:
        print("Ошибка при геренации файлa: ", str(e))
        exit()

if not data_list:
    print('Ошибка: не сгенерировано ни одной записи.')
    exit()


df = pd.DataFrame(data_list)
print('Данные сгенерированы, всего строк: ', str(len(df)))
try:
    df.to_excel('keyboard_sales.xlsx', index=False)
except Exception as e:
    print('Не удалось сохранить сгенерированный файл: ', str(e))
    exit()

#           анализ через пандас 
if not os.path.exists('keyboard_sales.xlsx'):
    print('Файл не найден')
    exit()

try:
    df = pd.read_excel('keyboard_sales.xlsx')
    print('Файл прочитан, строк: ', str(len(df)))
except Exception as e:
    print('Ошибка при чтении файла: ', str(e))

column = ['Цена', 'Количество', 'Товар']
if not all(col in df.columns for col in column):
    print('В файле нет нужных столбцов!!!')
    exit()

df['Цена'] = pd.to_numeric(df['Цена'], errors='coerce')
df['Количество'] = pd.to_numeric(df['Количество'], errors='coerce')
before = len(df)
df = df.dropna(subset=['Цена', 'Количество'])
after = len(df)
if before > after:
    if len(df) == 0:
        print('После очистки данных не осталось. Завершение.')
        exit()
    else: 
        print('Удалено ', before - after, ' строк с некорректными числами в файле')

# Добовляем столбец выручка
df['Выручка'] = df['Цена'] * df['Количество']
print('Добавлен столбец Выручка.')

try:
    svod = df.groupby('Товар')['Выручка'].sum().reset_index()
    print('Сводка по товарам готова.')
except Exception as e:
    print('Ошибка при группировке: ', str(e))

if len(svod) == 0:
    print('Свод по по товаром пуст. Завершение.')
    exit()

# берем топ 3 по выручке
sort = svod.sort_values('Выручка', ascending=False).head(3)

try:
    svod_data = df.groupby('Дата')['Выручка'].sum().reset_index()
    print('Сводка по дням готова.')
except Exception as e:
    print('Ошибка при группировке: ', str(e))

if len(svod_data) == 0:
    print('Свод по по датам пуст. Завершение.')
    exit()

#           --- Оформление ---

wb = Workbook()
ws = wb.active
print('Начинаю оформление отчёта...')

#       заголовок 

ws.merge_cells('A1:C1')
ws['A1'].value = 'Отчёт по клавиатурам и мышам'
ws['A1'].font = Font(bold=True, size=14)
ws['A1'].alignment = Alignment(horizontal='center')

#       под заголовок
ws.merge_cells('A2:B2')
ws['A2'].value = 'Топ-3 товаров'
ws['A2'].font = Font(bold=True)
ws['A2'].alignment = Alignment(horizontal='center')

#       стили

header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
header_font = Font(bold=True, color='FFFFFF')
center_alignment = Alignment(horizontal='center', vertical='center')
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

data = ['Товар', 'Выручка']

# шапка 
for col, v in enumerate(data, start=1):
    cell = ws.cell(row=3, column=col, value=v)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_alignment
    cell.border = border

# топ 3
if len(sort) == 0:
    print('Для таблицы "TOP 3" нехватает данных пропускаю...')
else:
    for row_idx, (_, row_data) in enumerate(sort.iterrows(), start=4):
        # row_data — это строка DataFrame, обращаемся по названиям колонок
        cell_a = ws.cell(row=row_idx, column=1, value=row_data['Товар'])
        cell_b = ws.cell(row=row_idx, column=2, value=row_data['Выручка'])
        for cell in (cell_a, cell_b):
          cell.alignment = center_alignment
          cell.border = border

# Выручка по дням

ws['A8'].value = 'Выручка по дням'
ws['A8'].font = Font(bold=True)
ws['A8'].alignment = Alignment(horizontal='center')

data2 = ['Дата', 'Выручка']

#       шапка
for col, v in enumerate(data2, start=1):
    cell = ws.cell(row=9, column=col, value=v)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_alignment
    cell.border = border

if len(svod_data) == 0:
    print('Для таблицы "Выручка по дням" нехватает данных пропускаю...')
else:
    for row, (_, row_data) in enumerate(svod_data.iterrows(), start=10):
        cell_a = ws.cell(row=row, column=1, value=row_data['Дата'])
        cell_b = ws.cell(row=row, column=2, value=row_data['Выручка'])
        for cell in (cell_a, cell_b):
            cell.alignment = center_alignment
            cell.border = border

# авто-подбор ширины

for col_diametr in ['A', 'B']:
    max_width = 0
    for cell in ws[col_diametr]:
        if cell.value:
            text_len = len(str(cell.value))
            if text_len > max_width:
                max_width = text_len
    ws.column_dimensions[col_diametr].width = max_width + 4

try:
    wb.save('отчёт_клавиатуры.xlsx')
    print('Файл успешно сохранен в отчёт_клавиатуры.xlsx')
except Exception as e:
    print('Ошибка при сохранении', str(e))
    exit()

print('Скрипт успешно завершил работу!')