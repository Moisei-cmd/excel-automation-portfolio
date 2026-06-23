import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

#               --- Стили ---

header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
header_font = Font(bold=True, color='FFFFFF')
center_alignment = Alignment(horizontal='center', vertical='center')
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

#               Функция авто-подбора ширины

def col_diametr():
    for col_diametr in ['A', 'B']:
        max = 0
        for cell in ws[col_diametr]:
            if cell.value:
                txt_len = len(str(cell.value))
                if txt_len > max:
                    max = txt_len
        ws.column_dimensions[col_diametr].width = max + 4

#       Запрос данных

file_path = input('Путь к Excel-файлу: ')
if not os.path.exists(file_path):
    print('Файл не найден!')
    exit()

col_price = input('Название столбца с ценой: ')
col_qty = input('Название столбца с количеством: ')
col_tovar = None
col_date = None
action = input('Что считаем? (1 - общая выручка, 2 - топ-5 товаров, 3 - выручка по дням): ')

#       Чтение
try:
    df = pd.read_excel(file_path)
    print('Файл прочитан, строк: ', str(len(df)))
except Exception as e:
    print('Ошибка при чтении файла: ', str(e))
    exit()

if action == '3':
    col_date = input('Название столбца с датой: ')
    column = [col_price, col_qty, col_date]
elif action == '2':
    col_tovar = input('Название столбца с товаром: ')
    column = [col_price, col_qty, col_tovar] 
else:
    column = [col_price, col_qty] 

if not all(col in df.columns for col in column):
    print('В файле нет нужных столбцов!!!')
    exit()

#       Очистка и вычисление выручки

df[col_price] = pd.to_numeric(df[col_price], errors='coerce')
df[col_qty] = pd.to_numeric(df[col_qty], errors='coerce')
before = len(df)
df = df.dropna(subset=[col_price, col_qty])
after = len(df)
if len(df) == 0:
    print('После очистки данных не осталось. Завершение.')
    exit()
if before > after:
    print('Удалено ', before - after, ' строк с некорректными числами в файле')

df['Выручка'] = df[col_price] * df[col_qty]

wb = Workbook()
ws = wb.active

#               --- main ---

if action == '1':
    total = df['Выручка'].sum()
    print('Общая выручка: ', str(total))
    data = ['Общая выручка', total]
    for i, col in enumerate(data, start=1):
        cell = ws.cell(row=1, column=i, value=col)
        cell.alignment = center_alignment
        cell.border = border

    col_diametr()

    wb.save('анализ_общая.xlsx')

elif action == '2':
    # 1) группировка и суммирование выручки
    grouped = df.groupby(col_tovar)['Выручка'].sum().reset_index()
    # 2) сортировка по выручке (по убыванию)
    top5 = grouped.sort_values('Выручка', ascending=False).head(5)
    ws.merge_cells('A1:B1')
    ws['A1'].value = "ТОП 5"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')

    headers = [col_tovar, 'Выручка']
    for col, value in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col, value=value)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = border

    for row_i, (_, row_data) in enumerate(top5.iterrows(), start=3):
        cell_a = ws.cell(row=row_i, column=1, value=row_data[col_tovar])
        cell_b = ws.cell(row=row_i, column=2, value=row_data['Выручка'])
        for cell in [cell_a, cell_b]:
            cell.alignment = center_alignment
            cell.border = border

    col_diametr()

    wb.save('анализ_топ5.xlsx')      

elif action == '3' :

    ws.merge_cells('A1:B1')
    ws['A1'].value = 'Выручка по дням'
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')

    headers = [col_date, 'Выручка']
    for col, value in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col, value=value)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = border

    sort_data = df.groupby(col_date)['Выручка'].sum().reset_index()
    sort_data = sort_data.sort_values(col_date)

    for row_x, (_, row_data) in enumerate(sort_data.iterrows(), start=3):
        cell_a = ws.cell(row=row_x, column=1, value=row_data[col_date])
        cell_b = ws.cell(row=row_x, column=2, value=row_data['Выручка'])
        for cell in [cell_a, cell_b]:
            cell.alignment = center_alignment
            cell.border = border
    col_diametr()

    wb.save('анализ_по_дням.xlsx')

