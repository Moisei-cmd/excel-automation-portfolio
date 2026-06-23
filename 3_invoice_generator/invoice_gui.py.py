import PySimpleGUI as sg
import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


# 1. Макет (layout) — список списков. Каждый внутренний список — строка в окне.
layout = [
    [sg.Text('Файл с данными (что продаем):')],
    [sg.Input(key='-Data-'), sg.FileBrowse()],
    [sg.Text('Файл шаблона счета:')],
    [sg.Input(key='-File-'), sg.FileBrowse()],
    [sg.Text('С какой строки вставлять товары (номер):')],
    [sg.Input(key='-Num_row-', default_text='5')],
    [sg.Text('Столбец с товаром: '), sg.Input(key='-Item-', default_text='Товар')],
    [sg.Text('Столбец с ценой:'), sg.Input(key='-Price-', default_text='Цена')],
    [sg.Text('Столбец с количеством:'), sg.Input(key='-Qty-', default_text='Количество')],
    [sg.Button('Создать счет'), sg.Button('Выход')],
    [sg.Multiline(size=(70, 12), key='-LOG-', autoscroll=True)]
]
window = sg.Window('Создание отчета', layout)


#           --- Стили ---

header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
header_font = Font(bold=True, color='FFFFFF')
center_alignment = Alignment(horizontal='center', vertical='center')
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

#           --- Функции ---

def generate():

    path_file = values['-Data-']
    path_example_file = values['-File-']
    if not os.path.exists(path_file):
        window['-LOG-'].update('Файл c данными не найден!\n', append=True)
        return
    if not os.path.exists(path_example_file):
        window['-LOG-'].update('Файл c примером не найден!\n', append=True)
        return

    try:
        df = pd.read_excel(path_file)
    except Exception as e:
        window['-LOG-'].update('Файл с данными не прочитан: ' + str(e) + '\n', append=True)
        return
    try:
        df_example = pd.read_excel(path_example_file)
    except Exception as e:
        window['-LOG-'].update('Файл с примером не прочитан:' + str(e) + '\n', append=True)
    

    tv = values['-Item-']
    price = values['-Price-']
    qty = values['-Qty-']

    column = [tv, price, qty]

    if not all(col in df.columns for col in column):
        window['-LOG-'].update('В файле нет нужных столбцов!!!\n', append=True)
        return


    df[price] = pd.to_numeric(df[price], errors='coerce')
    df[qty] = pd.to_numeric(df[qty], errors='coerce')
    before = len(df)
    df = df.dropna(subset=[price, qty])
    after = len(df)
    if len(df) == 0:
        window['-LOG-'].update('Файл пуст\n', append=True)
        return
    if before > after:
        window['-LOG-'].update('Удалено ' + str(before - after) + ' строк...\n', append=True)
    

    df['Сумма'] = df[price] * df[qty]

    wb = load_workbook(path_example_file)
    ws = wb.active
    try:
        start_row = int(values['-Num_row-'])
    except Exception as e:
        window['-LOG-'].update('Вставте число\n', append=True)
    column = [tv, price, qty, 'Сумма']
    for col, value in enumerate(column, start=1):
        cell = ws.cell(row=start_row, column=col, value=value)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = border

    data_start = start_row + 1
    for row_i, (_, row_data) in enumerate(df.iterrows(), start=data_start):
        cell_a = ws.cell(row=row_i, column=1, value=row_data[tv])
        cell_b = ws.cell(row=row_i, column=2, value=row_data[price])
        cell_c = ws.cell(row=row_i, column=3, value=row_data[qty])
        cell_d = ws.cell(row=row_i, column=4, value=row_data['Сумма'])
        for cell in [cell_a, cell_b, cell_c, cell_d]:
            cell.alignment = center_alignment
            cell.border = border
    

    total = df['Сумма'].sum()
    total_row = start_row + len(df) + 1  
    ws.cell(row=total_row, column=4, value='Итого:').font = Font(bold=True)
    ws.cell(row=total_row, column=5, value=total).font = Font(bold=True)

    
    save_path = path_example_file.replace('.xlsx', '_заполненный.xlsx')
    wb.save(save_path)
    window['-LOG-'].update('Счёт сохранён: ' + save_path + '\n', append=True)

while True:
    event, values = window.read()
    if event in (sg.WIN_CLOSED, 'Выход'):
        break
    if event == 'Создать счет':
        generate()

window.close()
